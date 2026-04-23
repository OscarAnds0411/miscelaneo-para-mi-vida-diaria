"""
extraer_catalogo.py  v1.0
--------------------------
Lee un .xlsx y extrae todos los valores únicos de MakeName y ModelName.
Genera catalogo_equivalencias.xlsx listo para que llenes la normalización.

Uso:
    python extraer_catalogo.py mi_base.xlsx
  O coloca el archivo en la misma carpeta y se detecta automáticamente.
"""

import os
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

BASE = os.path.dirname(os.path.abspath(__file__))

COL_MAKE  = "MakeName"
COL_MODEL = "ModelName"

# ── ESTILOS ────────────────────────────────────────────────────────────────────
def fill(color):   return PatternFill("solid", fgColor=color)
def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

C_MAKE_HDR  = "1F4E79"   # azul oscuro
C_MODEL_HDR = "1E5E38"   # verde oscuro
C_EVEN_M    = "D6E4F0"   # azul claro
C_EVEN_G    = "D5F0DC"   # verde claro
C_WHITE     = "FFFFFF"
C_YELLOW    = "FFF2CC"   # amarillo → columna a llenar


def hdr_cell(ws, row, col, value, bg):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = border()
    ws.row_dimensions[row].height = 24
    return c

def data_cell(ws, row, col, value, bg):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9)
    c.fill      = fill(bg)
    c.alignment = Alignment(vertical="center")
    c.border    = border()
    return c

def fill_cell(ws, row, col, bg=C_YELLOW):
    """Celda vacía amarilla → el usuario la llena."""
    c = ws.cell(row=row, column=col, value="")
    c.fill      = fill(bg)
    c.border    = border()
    c.font      = Font(name="Arial", size=9)
    return c


# ── MAIN ───────────────────────────────────────────────────────────────────────
def buscar_archivo():
    candidatos = [
        f for f in os.listdir(BASE)
        if f.lower().endswith(".xlsx")
        and not f.startswith("~")
        and "catalogo" not in f.lower()
    ]
    if not candidatos:
        return None
    candidatos.sort(key=lambda f: os.path.getmtime(os.path.join(BASE, f)), reverse=True)
    return os.path.join(BASE, candidatos[0])


def main():
    ruta = sys.argv[1] if len(sys.argv) > 1 else buscar_archivo()
    if not ruta or not os.path.exists(ruta):
        print("ERROR: No se encontró el archivo .xlsx")
        print("Uso: python extraer_catalogo.py mi_base.xlsx")
        sys.exit(1)

    print(f"Leyendo: {os.path.basename(ruta)}")
    df = pd.read_excel(ruta, dtype=str, engine="calamine")

    # Validar columnas
    for col in [COL_MAKE, COL_MODEL]:
        if col not in df.columns:
            print(f"ERROR: No se encontró la columna '{col}' en el archivo.")
            print(f"Columnas disponibles: {list(df.columns)}")
            sys.exit(1)

    makes  = sorted(df[COL_MAKE].dropna().str.strip().unique())
    models = sorted(df[COL_MODEL].dropna().str.strip().unique())

    print(f"  Makes únicos : {len(makes):,}")
    print(f"  Models únicos: {len(models):,}")

    # ── Construir Excel ────────────────────────────────────────────────────────
    wb  = Workbook()
    wb.remove(wb.active)

    # ── Hoja Makes ────────────────────────────────────────────────────────────
    ws_m = wb.create_sheet("Makes")
    ws_m.column_dimensions["A"].width = 28
    ws_m.column_dimensions["B"].width = 28
    ws_m.column_dimensions["C"].width = 18

    hdr_cell(ws_m, 1, 1, "MakeName (original)", C_MAKE_HDR)
    hdr_cell(ws_m, 1, 2, "MakeName (normalizado) ← LLENAR", C_MAKE_HDR)
    hdr_cell(ws_m, 1, 3, "# ocurrencias", C_MAKE_HDR)

    conteo_make = df[COL_MAKE].str.strip().value_counts()

    for r, make in enumerate(makes, start=2):
        bg = C_EVEN_M if r % 2 == 0 else C_WHITE
        data_cell(ws_m, r, 1, make, bg)
        fill_cell(ws_m, r, 2, C_YELLOW)
        data_cell(ws_m, r, 3, int(conteo_make.get(make, 0)), bg)

    ws_m.freeze_panes = "A2"
    ws_m.auto_filter.ref = f"A1:C{len(makes)+1}"

    # ── Hoja Models ───────────────────────────────────────────────────────────
    ws_g = wb.create_sheet("Models")
    ws_g.column_dimensions["A"].width = 30
    ws_g.column_dimensions["B"].width = 30
    ws_g.column_dimensions["C"].width = 28
    ws_g.column_dimensions["D"].width = 18

    hdr_cell(ws_g, 1, 1, "ModelName (original)", C_MODEL_HDR)
    hdr_cell(ws_g, 1, 2, "ModelName (normalizado) ← LLENAR", C_MODEL_HDR)
    hdr_cell(ws_g, 1, 3, "Makes asociados", C_MODEL_HDR)
    hdr_cell(ws_g, 1, 4, "# ocurrencias", C_MODEL_HDR)

    conteo_model = df[COL_MODEL].str.strip().value_counts()
    makes_por_model = (df.groupby(df[COL_MODEL].str.strip())[COL_MAKE]
                         .apply(lambda x: ", ".join(sorted(x.dropna().str.strip().unique())))
                         .to_dict())

    for r, model in enumerate(models, start=2):
        bg = C_EVEN_G if r % 2 == 0 else C_WHITE
        data_cell(ws_g, r, 1, model, bg)
        fill_cell(ws_g, r, 2, C_YELLOW)
        data_cell(ws_g, r, 3, makes_por_model.get(model, ""), bg)
        data_cell(ws_g, r, 4, int(conteo_model.get(model, 0)), bg)

    ws_g.freeze_panes = "A2"
    ws_g.auto_filter.ref = f"A1:D{len(models)+1}"

    # ── Hoja Instrucciones ────────────────────────────────────────────────────
    ws_i = wb.create_sheet("Instrucciones")
    ws_i.column_dimensions["A"].width = 70
    instrucciones = [
        ("INSTRUCCIONES DE USO", "1F4E79"),
        ("", None),
        ("1. Ve a la hoja 'Makes' y llena la columna B con el nombre normalizado.", None),
        ("   Ejemplo: VW → Volkswagen  |  CHEV → Chevrolet", None),
        ("   Si el valor ya está correcto, escribe lo mismo en la columna B.", None),
        ("", None),
        ("2. Ve a la hoja 'Models' y llena la columna B con el nombre normalizado.", None),
        ("   Ejemplo: JETTA → Jetta  |  SILVERADO → Silverado", None),
        ("   La columna C muestra qué Makes usan ese Model (contexto de referencia).", None),
        ("", None),
        ("3. Una vez llenado, corre el script 'aplicar_normalizacion.py' para", None),
        ("   aplicar los cambios a tu base completa automáticamente.", None),
        ("", None),
        ("TIPS:", "1E5E38"),
        ("  · Ordena por columna A para identificar variantes del mismo valor.", None),
        ("  · La columna '# ocurrencias' te dice qué valores son más críticos.", None),
        ("  · Las celdas amarillas son las que debes llenar.", None),
    ]
    for r, (texto, color) in enumerate(instrucciones, start=1):
        c = ws_i.cell(row=r, column=1, value=texto)
        if color:
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = fill(color)
        else:
            c.font = Font(name="Arial", size=10)
        ws_i.row_dimensions[r].height = 18

    # Mover Instrucciones al frente
    wb.move_sheet("Instrucciones", offset=-wb.index(wb["Instrucciones"]))

    salida = os.path.join(BASE, "catalogo_equivalencias.xlsx")
    wb.save(salida)
    print(f"\n✓ Guardado: catalogo_equivalencias.xlsx")
    print(f"  Hoja Makes : {len(makes)} valores únicos")
    print(f"  Hoja Models: {len(models)} valores únicos")
    print(f"\nLlena las columnas amarillas y avísame para generar el script de normalización.")


if __name__ == "__main__":
    main()
